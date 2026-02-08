"""
ENHANCED TENNIS BILLING - With Package Management, Coach Tracking & Mobile Optimization
Easy-to-use interface for tracking tennis lessons with advanced features
"""
from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from io import BytesIO
from sqlalchemy import extract, func
import calendar
import logging

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Import models
from models import db, Lesson, Coach, StudentPackage, Notification, MonthlyExport

# Initialize Flask app
app = Flask(__name__)

# Configuration
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'simple_billing.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'simple-tennis-billing-enhanced'

# Create exports directory
EXPORTS_DIR = os.path.join(basedir, 'exports')
if not os.path.exists(EXPORTS_DIR):
    os.makedirs(EXPORTS_DIR)

# Initialize database with app
db.init_app(app)


# ============================================================
# MAIN ROUTES
# ============================================================

@app.route('/')
def index():
    """Main page - Enhanced lesson entry with notifications"""
    today = date.today()
    
    # Get today's lessons
    today_lessons = Lesson.query.filter_by(date=today).order_by(Lesson.time).all()
    today_total = sum(lesson.invoice_amount for lesson in today_lessons)
    
    # Get this month's total
    month_start = today.replace(day=1)
    month_lessons = Lesson.query.filter(Lesson.date >= month_start).all()
    month_total = sum(lesson.invoice_amount for lesson in month_lessons)
    
    # Get active coaches
    coaches = Coach.query.filter_by(status='active').order_by(Coach.name).all()
    
    # Get unread notifications count
    unread_notifications = Notification.query.filter_by(is_read=False).count()
    
    # Get packages expiring soon (1 class remaining)
    expiring_packages = StudentPackage.query.filter(
        StudentPackage.status == 'active',
        StudentPackage.classes_remaining <= 1
    ).all()
    
    return render_template('simple_index.html', 
                         today_lessons=today_lessons,
                         today_total=today_total,
                         month_total=month_total,
                         today_date=today,
                         coaches=coaches,
                         unread_notifications=unread_notifications,
                         expiring_packages=expiring_packages)


@app.route('/api/lessons', methods=['POST'])
def add_lesson():
    """Add a new lesson with coach and package support"""
    data = request.get_json()
    
    # Create lesson
    lesson = Lesson(
        date=datetime.strptime(data['date'], '%Y-%m-%d').date(),
        time=data['time'],
        student_name=data['student_name'],
        hours=float(data['hours']),
        package=data.get('package', ''),
        invoice_amount=float(data['invoice_amount']),
        coach_id=data.get('coach_id'),
        package_id=data.get('package_id')
    )
    
    # Handle package deduction
    if data.get('package_id'):
        package = StudentPackage.query.get(data['package_id'])
        if package and package.deduct_class():
            lesson.deducted_from_package = True
            
            # Create notification if package is low
            if package.needs_renewal_alert():
                notification = Notification(
                    type='package_expiring',
                    title='Package Expiring Soon!',
                    message=f"{package.customer_name}'s package has {package.classes_remaining} class{'es' if package.classes_remaining != 1 else ''} remaining (${package.balance_remaining:.2f})",
                    priority='high',
                    student_name=package.customer_name,
                    package_id=package.id
                )
                db.session.add(notification)
    
    db.session.add(lesson)
    db.session.commit()
    
    return jsonify({'success': True, 'lesson': lesson.to_dict()})


@app.route('/api/lessons/<int:id>', methods=['DELETE'])
def delete_lesson(id):
    """Delete a lesson"""
    lesson = Lesson.query.get_or_404(id)
    db.session.delete(lesson)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/lessons/<int:id>', methods=['PUT'])
def update_lesson(id):
    """Update a lesson"""
    lesson = Lesson.query.get_or_404(id)
    data = request.get_json()
    
    lesson.date = datetime.strptime(data['date'], '%Y-%m-%d').date()
    lesson.time = data['time']
    lesson.student_name = data['student_name']
    lesson.hours = float(data['hours'])
    lesson.package = data.get('package', '')
    lesson.invoice_amount = float(data['invoice_amount'])
    
    db.session.commit()
    return jsonify({'success': True, 'lesson': lesson.to_dict()})


@app.route('/history')
def history():
    """View all lessons history"""
    selected_date = request.args.get('date')
    
    try:
        if selected_date:
            filter_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
            lessons = Lesson.query.filter_by(date=filter_date).order_by(Lesson.time).all()
        else:
            lessons = Lesson.query.order_by(Lesson.date.desc(), Lesson.time).limit(100).all()
        
        total = sum(lesson.invoice_amount for lesson in lessons)
        unread_notifications = Notification.query.filter_by(is_read=False).count()
        
        return render_template('simple_history.html', lessons=lessons, total=total, selected_date=selected_date, unread_notifications=unread_notifications)
    except ValueError:
        # Invalid date format
        lessons = Lesson.query.order_by(Lesson.date.desc(), Lesson.time).limit(100).all()
        total = sum(lesson.invoice_amount for lesson in lessons)
        unread_notifications = Notification.query.filter_by(is_read=False).count()
        return render_template('simple_history.html', lessons=lessons, total=total, selected_date=None, unread_notifications=unread_notifications)


# ============================================================
# PACKAGE MANAGEMENT API
# ============================================================

@app.route('/packages')
def packages_page():
    """Packages management page"""
    packages = StudentPackage.query.order_by(
        StudentPackage.status.desc(),
        StudentPackage.classes_remaining
    ).all()
    unread_notifications = Notification.query.filter_by(is_read=False).count()
    return render_template('packages.html', packages=packages, unread_notifications=unread_notifications)


@app.route('/api/packages', methods=['GET'])
def get_packages():
    """Get all packages"""
    status_filter = request.args.get('status', 'all')
    
    query = StudentPackage.query
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    packages = query.order_by(
        StudentPackage.status.desc(),
        StudentPackage.classes_remaining
    ).all()
    
    return jsonify({
        'success': True,
        'packages': [p.to_dict() for p in packages]
    })


@app.route('/api/packages/<int:id>', methods=['GET'])
def get_package(id):
    """Get package details"""
    package = StudentPackage.query.get_or_404(id)
    return jsonify({
        'success': True,
        'package': package.to_dict()
    })


@app.route('/api/packages', methods=['POST'])
def create_package():
    """Create a new package"""
    data = request.get_json()
    
    total_amount = float(data['total_amount'])
    amount_per_class = float(data['amount_per_class'])
    total_classes = int(total_amount / amount_per_class)
    
    package = StudentPackage(
        customer_name=data['customer_name'],
        package_name=data['package_name'],
        total_amount=total_amount,
        amount_per_class=amount_per_class,
        total_classes=total_classes,
        classes_remaining=total_classes,
        balance_remaining=total_amount,
        start_date=datetime.strptime(data['start_date'], '%Y-%m-%d').date() if data.get('start_date') else date.today(),
        expiry_date=datetime.strptime(data['expiry_date'], '%Y-%m-%d').date() if data.get('expiry_date') else None
    )
    
    db.session.add(package)
    db.session.commit()
    
    return jsonify({'success': True, 'package': package.to_dict()})


@app.route('/api/packages/<int:id>', methods=['PUT'])
def update_package(id):
    """Update package details"""
    package = StudentPackage.query.get_or_404(id)
    data = request.get_json()
    
    package.customer_name = data.get('customer_name', package.customer_name)
    package.package_name = data.get('package_name', package.package_name)
    package.status = data.get('status', package.status)
    
    if data.get('expiry_date'):
        package.expiry_date = datetime.strptime(data['expiry_date'], '%Y-%m-%d').date()
    
    db.session.commit()
    
    return jsonify({'success': True, 'package': package.to_dict()})


@app.route('/api/packages/<int:id>', methods=['DELETE'])
def delete_package(id):
    """Delete a package"""
    package = StudentPackage.query.get_or_404(id)
    db.session.delete(package)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/packages/check/<customer_name>')
def check_package(customer_name):
    """Check if customer has active package"""
    try:
        # Try exact match first
        package = StudentPackage.query.filter_by(
            customer_name=customer_name,
            status='active'
        ).filter(StudentPackage.classes_remaining > 0).first()
        
        # If no exact match, try case-insensitive
        if not package:
            package = StudentPackage.query.filter(
                db.func.lower(StudentPackage.customer_name) == customer_name.lower(),
                StudentPackage.status == 'active',
                StudentPackage.classes_remaining > 0
            ).first()
        
        if package:
            return jsonify({
                'success': True,
                'has_package': True,
                'package': package.to_dict()
            })
        else:
            return jsonify({
                'success': True,
                'has_package': False
            })
    except Exception as e:
        logger.error(f"Error checking package: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


# ============================================================
# COACH MANAGEMENT API
# ============================================================

@app.route('/coaches')
def coaches_page():
    """Coaches management page"""
    coaches = Coach.query.filter_by(status='active').order_by(Coach.name).all()
    
    # Get current month stats for each coach
    today = date.today()
    for coach in coaches:
        coach.monthly_hours = coach.get_monthly_hours(today.year, today.month)
        coach.monthly_lessons = len([l for l in coach.lessons if l.date.year == today.year and l.date.month == today.month])
        coach.monthly_earnings = coach.monthly_hours * coach.hourly_rate
    
    unread_notifications = Notification.query.filter_by(is_read=False).count()
    return render_template('coaches.html', coaches=coaches, current_month=today.strftime('%B %Y'), unread_notifications=unread_notifications)


@app.route('/api/coaches', methods=['GET'])
def get_coaches():
    """Get all coaches"""
    status_filter = request.args.get('status', 'active')
    
    query = Coach.query
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    coaches = query.order_by(Coach.name).all()
    
    return jsonify({
        'success': True,
        'coaches': [c.to_dict() for c in coaches]
    })


@app.route('/api/coaches/<int:id>', methods=['GET'])
def get_coach(id):
    """Get coach details"""
    coach = Coach.query.get_or_404(id)
    return jsonify({
        'success': True,
        'coach': coach.to_dict()
    })


@app.route('/api/coaches', methods=['POST'])
def create_coach():
    """Create a new coach"""
    data = request.get_json()
    
    coach = Coach(
        name=data['name'],
        email=data.get('email', ''),
        phone=data.get('phone', ''),
        hourly_rate=float(data.get('hourly_rate', 0)),
        status=data.get('status', 'active')
    )
    
    db.session.add(coach)
    db.session.commit()
    
    return jsonify({'success': True, 'coach': coach.to_dict()})


@app.route('/api/coaches/<int:id>', methods=['PUT'])
def update_coach(id):
    """Update coach details"""
    coach = Coach.query.get_or_404(id)
    data = request.get_json()
    
    coach.name = data.get('name', coach.name)
    coach.email = data.get('email', coach.email)
    coach.phone = data.get('phone', coach.phone)
    coach.hourly_rate = float(data.get('hourly_rate', coach.hourly_rate))
    coach.status = data.get('status', coach.status)
    
    db.session.commit()
    
    return jsonify({'success': True, 'coach': coach.to_dict()})


@app.route('/api/coaches/<int:id>', methods=['DELETE'])
def delete_coach(id):
    """Delete a coach"""
    coach = Coach.query.get_or_404(id)
    db.session.delete(coach)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/coaches/<int:id>/breakdown')
def coach_breakdown(id):
    """Get daily breakdown for a coach"""
    coach = Coach.query.get_or_404(id)
    
    year = request.args.get('year', date.today().year, type=int)
    month = request.args.get('month', date.today().month, type=int)
    
    breakdown = coach.get_daily_breakdown(year, month)
    
    return jsonify({
        'success': True,
        'coach': coach.to_dict(),
        'year': year,
        'month': month,
        'month_name': calendar.month_name[month],
        'breakdown': breakdown
    })


# ============================================================
# NOTIFICATION API
# ============================================================

@app.route('/notifications')
def notifications_page():
    """Notifications page"""
    notifications = Notification.query.order_by(
        Notification.is_read,
        Notification.created_at.desc()
    ).limit(50).all()
    unread_notifications = Notification.query.filter_by(is_read=False).count()
    return render_template('notifications.html', notifications=notifications, unread_notifications=unread_notifications)


@app.route('/api/notifications', methods=['GET'])
def get_notifications():
    """Get all notifications"""
    unread_only = request.args.get('unread_only', 'false').lower() == 'true'
    
    query = Notification.query
    if unread_only:
        query = query.filter_by(is_read=False)
    
    notifications = query.order_by(
        Notification.is_read,
        Notification.created_at.desc()
    ).limit(50).all()
    
    return jsonify({
        'success': True,
        'notifications': [n.to_dict() for n in notifications],
        'unread_count': Notification.query.filter_by(is_read=False).count()
    })


@app.route('/api/notifications/<int:id>/mark_read', methods=['POST'])
def mark_notification_read(id):
    """Mark notification as read"""
    notification = Notification.query.get_or_404(id)
    notification.is_read = True
    db.session.commit()
    
    return jsonify({'success': True})


@app.route('/api/notifications/mark_all_read', methods=['POST'])
def mark_all_notifications_read():
    """Mark all notifications as read"""
    Notification.query.filter_by(is_read=False).update({'is_read': True})
    db.session.commit()
    
    return jsonify({'success': True})


@app.route('/api/notifications/<int:id>', methods=['DELETE'])
def delete_notification(id):
    """Delete a notification"""
    notification = Notification.query.get_or_404(id)
    db.session.delete(notification)
    db.session.commit()
    return jsonify({'success': True})


# ============================================================
# STUDENT NAME AUTOCOMPLETE
# ============================================================

@app.route('/api/students/search')
def search_students():
    """Search for student names (for autocomplete)"""
    query = request.args.get('q', '').strip()
    
    if len(query) < 2:
        return jsonify({'success': True, 'students': []})
    
    # Get unique student names from lessons
    students = db.session.query(Lesson.student_name).distinct()\
        .filter(Lesson.student_name.ilike(f'%{query}%'))\
        .order_by(Lesson.student_name)\
        .limit(10).all()
    
    student_names = [s[0] for s in students]
    
    return jsonify({
        'success': True,
        'students': student_names
    })


# ============================================================
# IMPORT/EXPORT ROUTES
# ============================================================

@app.route('/import_excel', methods=['POST'])
def import_excel():
    """Import lessons from Excel file"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'})
    
    try:
        # Read Excel file
        wb = openpyxl.load_workbook(file)
        sheet = wb.active
        
        imported_count = 0
        
        # Skip header row, start from row 2
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip if no date
                continue
            
            try:
                lesson_date = row[0] if isinstance(row[0], date) else datetime.strptime(str(row[0]), '%Y-%m-%d').date()
                time_str = str(row[1]) if row[1] else ''
                student_name = str(row[2]) if row[2] else 'Unknown'
                hours = float(row[3]) if row[3] else 1.0
                package = str(row[4]) if row[4] else ''
                invoice = float(row[5]) if row[5] else 0.0
                
                # Check if lesson already exists
                existing = Lesson.query.filter_by(
                    date=lesson_date,
                    time=time_str,
                    student_name=student_name
                ).first()
                
                if not existing:
                    lesson = Lesson(
                        date=lesson_date,
                        time=time_str,
                        student_name=student_name,
                        hours=hours,
                        package=package,
                        invoice_amount=invoice
                    )
                    db.session.add(lesson)
                    imported_count += 1
            except Exception as e:
                print(f"Error importing row: {e}")
                continue
        
        db.session.commit()
        return jsonify({'success': True, 'imported': imported_count})
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/export_excel')
def export_excel():
    """Export all lessons to Excel file"""
    # Get all lessons
    lessons = Lesson.query.order_by(Lesson.date, Lesson.time).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Tennis Lessons"
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Write headers
    headers = ['Date', 'Time', "Student's Name/Program", 'Hours', 'Coach', 'Package/Group', 'Invoice']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row, lesson in enumerate(lessons, 2):
        sheet.cell(row=row, column=1, value=lesson.date)
        sheet.cell(row=row, column=2, value=lesson.time)
        sheet.cell(row=row, column=3, value=lesson.student_name)
        sheet.cell(row=row, column=4, value=lesson.hours)
        sheet.cell(row=row, column=5, value=lesson.coach.name if lesson.coach else '')
        sheet.cell(row=row, column=6, value=lesson.package)
        sheet.cell(row=row, column=7, value=lesson.invoice_amount)
    
    # Adjust column widths
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 8
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 10
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Tennis_Lessons_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ============================================================
# MONTHLY EXPORT API
# ============================================================

@app.route('/api/export_monthly/<int:year>/<int:month>')
def export_monthly(year, month):
    """Generate monthly export Excel file"""
    month_name = calendar.month_name[month]
    
    # Get lessons for the month
    lessons = Lesson.query.filter(
        extract('year', Lesson.date) == year,
        extract('month', Lesson.date) == month
    ).order_by(Lesson.date, Lesson.time).all()
    
    # Create workbook with multiple sheets
    wb = openpyxl.Workbook()
    
    # Sheet 1: Lessons Summary
    sheet1 = wb.active
    sheet1.title = "Lessons Summary"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    headers = ['Date', 'Time', 'Student', 'Hours', 'Coach', 'Package', 'Amount']
    for col, header in enumerate(headers, 1):
        cell = sheet1.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row, lesson in enumerate(lessons, 2):
        sheet1.cell(row=row, column=1, value=lesson.date)
        sheet1.cell(row=row, column=2, value=lesson.time)
        sheet1.cell(row=row, column=3, value=lesson.student_name)
        sheet1.cell(row=row, column=4, value=lesson.hours)
        sheet1.cell(row=row, column=5, value=lesson.coach.name if lesson.coach else '')
        sheet1.cell(row=row, column=6, value=lesson.package)
        sheet1.cell(row=row, column=7, value=lesson.invoice_amount)
    
    # Sheet 2: Daily Totals
    sheet2 = wb.create_sheet("Daily Totals")
    daily_totals = {}
    for lesson in lessons:
        date_key = lesson.date
        if date_key not in daily_totals:
            daily_totals[date_key] = {'lessons': 0, 'hours': 0, 'revenue': 0}
        daily_totals[date_key]['lessons'] += 1
        daily_totals[date_key]['hours'] += lesson.hours
        daily_totals[date_key]['revenue'] += lesson.invoice_amount
    
    headers2 = ['Date', 'Lessons', 'Hours', 'Revenue']
    for col, header in enumerate(headers2, 1):
        cell = sheet2.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row, (date_key, totals) in enumerate(sorted(daily_totals.items()), 2):
        sheet2.cell(row=row, column=1, value=date_key)
        sheet2.cell(row=row, column=2, value=totals['lessons'])
        sheet2.cell(row=row, column=3, value=totals['hours'])
        sheet2.cell(row=row, column=4, value=totals['revenue'])
    
    # Sheet 3: Coach Hours
    sheet3 = wb.create_sheet("Coach Hours")
    coaches = Coach.query.filter_by(status='active').all()
    
    headers3 = ['Coach', 'Total Hours', 'Lessons', 'Hourly Rate', 'Earnings']
    for col, header in enumerate(headers3, 1):
        cell = sheet3.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row, coach in enumerate(coaches, 2):
        coach_lessons = [l for l in lessons if l.coach_id == coach.id]
        total_hours = sum(l.hours for l in coach_lessons)
        earnings = total_hours * coach.hourly_rate
        
        sheet3.cell(row=row, column=1, value=coach.name)
        sheet3.cell(row=row, column=2, value=total_hours)
        sheet3.cell(row=row, column=3, value=len(coach_lessons))
        sheet3.cell(row=row, column=4, value=coach.hourly_rate)
        sheet3.cell(row=row, column=5, value=earnings)
    
    # Sheet 4: Monthly Summary
    sheet4 = wb.create_sheet("Monthly Summary")
    total_lessons = len(lessons)
    total_hours = sum(l.hours for l in lessons)
    total_revenue = sum(l.invoice_amount for l in lessons)
    
    summary_data = [
        ['Month', f"{month_name} {year}"],
        ['Total Lessons', total_lessons],
        ['Total Hours', total_hours],
        ['Total Revenue', f"${total_revenue:.2f}"],
        ['Average per Lesson', f"${total_revenue/total_lessons:.2f}" if total_lessons > 0 else "$0.00"],
        ['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    for row, (label, value) in enumerate(summary_data, 1):
        sheet4.cell(row=row, column=1, value=label).font = Font(bold=True)
        sheet4.cell(row=row, column=2, value=value)
    
    # Adjust column widths for all sheets
    for sheet in [sheet1, sheet2, sheet3, sheet4]:
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to file
    filename = f"{month_name}_{year}.xlsx"
    filepath = os.path.join(EXPORTS_DIR, filename)
    wb.save(filepath)
    
    # Save export record
    export_record = MonthlyExport(
        month=month,
        year=year,
        month_name=month_name,
        filename=filename,
        file_path=filepath,
        total_lessons=total_lessons,
        total_revenue=total_revenue,
        total_hours=total_hours,
        exported_by='manual'
    )
    db.session.add(export_record)
    
    # Create notification
    notification = Notification(
        type='monthly_export',
        title='Monthly Export Complete',
        message=f"Excel report for {month_name} {year} has been generated successfully. Total: ${total_revenue:.2f}",
        priority='medium'
    )
    db.session.add(notification)
    
    db.session.commit()
    
    # Send file
    return send_file(
        filepath,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/exports')
def get_exports():
    """Get all monthly exports"""
    exports = MonthlyExport.query.order_by(
        MonthlyExport.year.desc(),
        MonthlyExport.month.desc()
    ).all()
    
    return jsonify({
        'success': True,
        'exports': [e.to_dict() for e in exports]
    })


# ============================================================
# OFFLINE SUPPORT - Service Worker & Manifest
# ============================================================

@app.route('/sw.js')
def service_worker():
    """Serve service worker for offline support"""
    return app.send_static_file('sw.js'), 200, {'Content-Type': 'application/javascript'}


@app.route('/manifest.json')
def manifest():
    """Serve PWA manifest"""
    return app.send_static_file('manifest.json'), 200, {'Content-Type': 'application/json'}


# ============================================================
# ANALYTICS API
# ============================================================

@app.route('/analytics')
def analytics_page():
    """Analytics dashboard page"""
    unread_notifications = Notification.query.filter_by(is_read=False).count()
    return render_template('analytics.html', unread_notifications=unread_notifications)


@app.route('/api/analytics')
def get_analytics():
    """Get comprehensive analytics data"""
    today = date.today()
    month_start = today.replace(day=1)
    
    # Metrics
    total_lessons = Lesson.query.count()
    total_revenue = db.session.query(func.sum(Lesson.invoice_amount)).scalar() or 0.0
    active_packages = StudentPackage.query.filter_by(status='active').count()
    active_coaches = Coach.query.filter_by(status='active').count()
    
    # Revenue trend (last 30 days)
    revenue_trend_labels = []
    revenue_trend_data = []
    for i in range(29, -1, -1):
        day = today - timedelta(days=i)
        day_lessons = Lesson.query.filter_by(date=day).all()
        day_revenue = sum(l.invoice_amount for l in day_lessons)
        revenue_trend_labels.append(day.strftime('%m/%d'))
        revenue_trend_data.append(float(day_revenue))
    
    # Coach revenue
    coaches = Coach.query.filter_by(status='active').all()
    coach_revenue_labels = []
    coach_revenue_data = []
    for coach in coaches:
        coach_lessons = [l for l in coach.lessons if l.date >= month_start]
        coach_total = sum(l.invoice_amount for l in coach_lessons)
        if coach_total > 0:
            coach_revenue_labels.append(coach.name)
            coach_revenue_data.append(float(coach_total))
    
    # Lessons per day (this month)
    lessons_labels = []
    lessons_data = []
    for i in range(1, today.day + 1):
        day = today.replace(day=i)
        day_count = Lesson.query.filter_by(date=day).count()
        lessons_labels.append(str(i))
        lessons_data.append(day_count)
    
    # Package status
    active_count = StudentPackage.query.filter_by(status='active').count()
    depleted_count = StudentPackage.query.filter_by(status='depleted').count()
    expired_count = StudentPackage.query.filter_by(status='expired').count()
    
    # Top students
    student_stats = db.session.query(
        Lesson.student_name,
        func.count(Lesson.id).label('lessons'),
        func.sum(Lesson.invoice_amount).label('revenue')
    ).group_by(Lesson.student_name).order_by(func.sum(Lesson.invoice_amount).desc()).limit(5).all()
    
    top_students = [{
        'name': s.student_name,
        'lessons': s.lessons,
        'revenue': float(s.revenue or 0)
    } for s in student_stats]
    
    # Top coaches
    top_coaches = []
    for coach in coaches:
        coach_lessons = [l for l in coach.lessons if l.date >= month_start]
        total_hours = sum(l.hours for l in coach_lessons)
        earnings = total_hours * coach.hourly_rate
        if earnings > 0:
            top_coaches.append({
                'name': coach.name,
                'hours': total_hours,
                'earnings': earnings
            })
    top_coaches.sort(key=lambda x: x['earnings'], reverse=True)
    top_coaches = top_coaches[:5]
    
    return jsonify({
        'success': True,
        'metrics': {
            'total_revenue': float(total_revenue),
            'total_lessons': total_lessons,
            'active_packages': active_packages,
            'active_coaches': active_coaches
        },
        'charts': {
            'revenue_trend': {
                'labels': revenue_trend_labels,
                'data': revenue_trend_data
            },
            'coach_revenue': {
                'labels': coach_revenue_labels,
                'data': coach_revenue_data
            },
            'lessons_per_day': {
                'labels': lessons_labels,
                'data': lessons_data
            },
            'package_status': {
                'labels': ['Active', 'Depleted', 'Expired'],
                'data': [active_count, depleted_count, expired_count]
            }
        },
        'top_students': top_students,
        'top_coaches': top_coaches
    })


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        
    # Initialize scheduler for monthly exports
    from scheduler import init_scheduler
    scheduler = init_scheduler(app, db, Lesson, Coach, MonthlyExport, Notification, EXPORTS_DIR)
    
    # Run on all interfaces to be accessible from other devices
    app.run(host='0.0.0.0', port=5001, debug=True)
