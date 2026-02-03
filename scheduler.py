"""
Scheduler for automated monthly Excel exports
Runs at 11:59 PM on the last day of each month
"""
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from datetime import datetime, date
import calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
from sqlalchemy import extract


def create_monthly_export(app, db, Lesson, Coach, MonthlyExport, Notification, EXPORTS_DIR):
    """
    Generate monthly export automatically
    Called by scheduler at end of month
    """
    with app.app_context():
        # Get previous month (since this runs at end of month)
        today = date.today()
        year = today.year
        month = today.month
        
        # If it's January, export December of previous year
        if month == 1:
            month = 12
            year = year - 1
        
        month_name = calendar.month_name[month]
        
        print(f"Starting automatic monthly export for {month_name} {year}")
        
        # Get lessons for the month
        lessons = Lesson.query.filter(
            extract('year', Lesson.date) == year,
            extract('month', Lesson.date) == month
        ).order_by(Lesson.date, Lesson.time).all()
        
        if not lessons:
            print(f"No lessons found for {month_name} {year}, skipping export")
            return
        
        # Create workbook with multiple sheets
        wb = openpyxl.Workbook()
        
        # Sheet 1: Lessons Summary
        sheet1 = wb.active
        sheet1.title = "Lessons Summary"
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        headers = ['Date', 'Time', 'Student', 'Hours', 'Coach', 'Package', 'Amount']
        for col, header in enumerate(headers, 1):
            cell = sheet1.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row, lesson in enumerate(lessons, 2):
            sheet1.cell(row=row, column=1, value=lesson.date)
            sheet1.cell(row=row, column=2, value=lesson.time)
            sheet1.cell(row=row, column=3, value=lesson.student_name)
            sheet1.cell(row=row, column=4, value=lesson.hours)
            sheet1.cell(row=row, column=5, value=lesson.coach.name if lesson.coach else '')
            sheet1.cell(row=row, column=6, value=lesson.package)
            sheet1.cell(row=row, column=7, value=lesson.invoice_amount)
        
        # Sheet 2: Daily Totals
        sheet2 = wb.create_sheet("Daily Totals")
        daily_totals = {}
        for lesson in lessons:
            date_key = lesson.date
            if date_key not in daily_totals:
                daily_totals[date_key] = {'lessons': 0, 'hours': 0, 'revenue': 0}
            daily_totals[date_key]['lessons'] += 1
            daily_totals[date_key]['hours'] += lesson.hours
            daily_totals[date_key]['revenue'] += lesson.invoice_amount
        
        headers2 = ['Date', 'Lessons', 'Hours', 'Revenue']
        for col, header in enumerate(headers2, 1):
            cell = sheet2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row, (date_key, totals) in enumerate(sorted(daily_totals.items()), 2):
            sheet2.cell(row=row, column=1, value=date_key)
            sheet2.cell(row=row, column=2, value=totals['lessons'])
            sheet2.cell(row=row, column=3, value=totals['hours'])
            sheet2.cell(row=row, column=4, value=totals['revenue'])
        
        # Sheet 3: Coach Hours
        sheet3 = wb.create_sheet("Coach Hours")
        coaches = Coach.query.filter_by(status='active').all()
        
        headers3 = ['Coach', 'Total Hours', 'Lessons', 'Hourly Rate', 'Earnings']
        for col, header in enumerate(headers3, 1):
            cell = sheet3.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row, coach in enumerate(coaches, 2):
            coach_lessons = [l for l in lessons if l.coach_id == coach.id]
            total_hours = sum(l.hours for l in coach_lessons)
            earnings = total_hours * coach.hourly_rate
            
            sheet3.cell(row=row, column=1, value=coach.name)
            sheet3.cell(row=row, column=2, value=total_hours)
            sheet3.cell(row=row, column=3, value=len(coach_lessons))
            sheet3.cell(row=row, column=4, value=coach.hourly_rate)
            sheet3.cell(row=row, column=5, value=earnings)
        
        # Sheet 4: Package Usage
        sheet4 = wb.create_sheet("Package Usage")
        package_lessons = [l for l in lessons if l.deducted_from_package]
        
        headers4 = ['Date', 'Student', 'Package Used', 'Amount']
        for col, header in enumerate(headers4, 1):
            cell = sheet4.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row, lesson in enumerate(package_lessons, 2):
            sheet4.cell(row=row, column=1, value=lesson.date)
            sheet4.cell(row=row, column=2, value=lesson.student_name)
            sheet4.cell(row=row, column=3, value=lesson.package)
            sheet4.cell(row=row, column=4, value=lesson.invoice_amount)
        
        # Sheet 5: Monthly Summary
        sheet5 = wb.create_sheet("Monthly Summary")
        total_lessons = len(lessons)
        total_hours = sum(l.hours for l in lessons)
        total_revenue = sum(l.invoice_amount for l in lessons)
        package_revenue = sum(l.invoice_amount for l in package_lessons)
        
        summary_data = [
            ['Month', f"{month_name} {year}"],
            ['Total Lessons', total_lessons],
            ['Total Hours', total_hours],
            ['Total Revenue', f"${total_revenue:.2f}"],
            ['Package Revenue', f"${package_revenue:.2f}"],
            ['Regular Revenue', f"${total_revenue - package_revenue:.2f}"],
            ['Average per Lesson', f"${total_revenue/total_lessons:.2f}" if total_lessons > 0 else "$0.00"],
            ['Export Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Export Type', 'Automatic']
        ]
        
        for row, (label, value) in enumerate(summary_data, 1):
            sheet5.cell(row=row, column=1, value=label).font = Font(bold=True, size=14)
            sheet5.cell(row=row, column=2, value=value).font = Font(size=14)
        
        # Adjust column widths for all sheets
        for sheet in [sheet1, sheet2, sheet3, sheet4, sheet5]:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to file
        filename = f"{month_name}_{year}.xlsx"
        filepath = os.path.join(EXPORTS_DIR, filename)
        wb.save(filepath)
        
        # Save export record
        export_record = MonthlyExport(
            month=month,
            year=year,
            month_name=month_name,
            filename=filename,
            file_path=filepath,
            total_lessons=total_lessons,
            total_revenue=total_revenue,
            total_hours=total_hours,
            exported_by='system'
        )
        db.session.add(export_record)
        
        # Create notification
        notification = Notification(
            type='monthly_export',
            title='Monthly Export Complete',
            message=f"Automatic export for {month_name} {year} completed successfully. {total_lessons} lessons, ${total_revenue:.2f} total revenue.",
            priority='medium'
        )
        db.session.add(notification)
        
        db.session.commit()
        
        print(f"Monthly export completed: {filename}")
        print(f"Total: {total_lessons} lessons, ${total_revenue:.2f}")


def init_scheduler(app, db, Lesson, Coach, MonthlyExport, Notification, EXPORTS_DIR):
    """
    Initialize the scheduler for monthly exports
    Runs at 11:59 PM on the last day of each month
    """
    scheduler = BackgroundScheduler()
    
    # Schedule monthly export - runs at 11:59 PM on last day of month
    # Cron: minute=59, hour=23, day=last day of month
    trigger = CronTrigger(
        day='last',  # Last day of month
        hour=23,     # 11 PM
        minute=59    # 59 minutes
    )
    
    scheduler.add_job(
        func=lambda: create_monthly_export(app, db, Lesson, Coach, MonthlyExport, Notification, EXPORTS_DIR),
        trigger=trigger,
        id='monthly_export',
        name='Generate Monthly Excel Report',
        replace_existing=True
    )
    
    # Optional: Add a test job that runs daily at 11:50 PM (for testing)
    # Uncomment the following lines if you want to test the export daily
    # test_trigger = CronTrigger(hour=23, minute=50)
    # scheduler.add_job(
    #     func=lambda: create_monthly_export(app, db, Lesson, Coach, MonthlyExport, Notification, EXPORTS_DIR),
    #     trigger=test_trigger,
    #     id='test_monthly_export',
    #     name='Test Monthly Export (Daily)',
    #     replace_existing=True
    # )
    
    scheduler.start()
    print("Monthly export scheduler started!")
    print("   - Automatic exports will run at 11:59 PM on the last day of each month")
    
    return scheduler
